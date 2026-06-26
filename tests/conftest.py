import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytest


@pytest.fixture
def build_simple_xlsx(tmp_path):
    def _build(name="simple.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spec"
        ws["A1"] = "Parameter"; ws["B1"] = "280Ah"; ws["C1"] = "314Ah"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = "Loading Level"; ws["B2"] = 25.40; ws["C2"] = 38.20
        ws["B2"].number_format = "0.00"; ws["C2"].number_format = "0.00"
        ws["A3"] = "N/P ratio"; ws["B3"] = 1.087; ws["C3"] = 1.086
        ws["B3"].number_format = "0.000"; ws["C3"].number_format = "0.000"
        ws.column_dimensions["A"].width = 18
        p = tmp_path / name
        wb.save(p)
        return str(p)
    return _build


@pytest.fixture
def build_rich_xlsx(tmp_path):
    """Merged header, a fill, a percent, a blank cell, and an embedded image."""
    def _build(name="rich.xlsx", with_image=True):
        from PIL import Image
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cell Design"
        ws.merge_cells("A1:C1")
        ws["A1"] = "ESS LFP Cell Design"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill("solid", fgColor="FFD966")
        ws["A2"] = "Capacity"; ws["B2"] = "280Ah"; ws["C2"] = "314Ah"
        ws["A3"] = "Margin"; ws["B3"] = 0.012; ws["B3"].number_format = "0.0%"
        ws["A4"] = "Note"; ws["C4"] = None  # intentional blank
        if with_image:
            img_path = tmp_path / "_e.png"
            Image.new("RGB", (40, 20), (10, 80, 160)).save(img_path)
            from openpyxl.drawing.image import Image as XLImage
            ws.add_image(XLImage(str(img_path)), "E2")
        p = tmp_path / name
        wb.save(p)
        return str(p)
    return _build
