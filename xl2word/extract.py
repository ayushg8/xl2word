from __future__ import annotations
import os
import shutil
import zipfile
import openpyxl
from .model import Workbook, Sheet, Cell, Style, MergedRange, ImageAsset
from .cleaners import format_cell_value


def _rgb(color) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return None
    if len(rgb) == 8:
        if rgb == "00000000":        # no-color sentinel
            return None
        rgb = rgb[2:]                # strip ARGB alpha prefix
    return rgb.upper()


def _cell_style(c) -> Style:
    fill = None
    if c.fill is not None and c.fill.patternType:
        fill = _rgb(c.fill.fgColor)
    border = False
    if c.border is not None:
        border = any(getattr(c.border, side).style
                     for side in ("top", "bottom", "left", "right"))
    return Style(
        bold=bool(c.font and c.font.bold),
        italic=bool(c.font and c.font.italic),
        font_name=c.font.name if c.font else None,
        font_size=float(c.font.size) if c.font and c.font.size else None,
        font_color=_rgb(c.font.color) if c.font else None,
        fill=fill,
        align_h=c.alignment.horizontal if c.alignment else None,
        align_v=c.alignment.vertical if c.alignment else None,
        border=border,
        number_format=c.number_format,
    )


def extract_semantic(xlsx_path: str) -> Workbook:
    book = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets: list[Sheet] = []
    for idx, ws in enumerate(book.worksheets):
        cells: list[Cell] = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is None and (c.fill is None or not c.fill.patternType):
                    continue  # skip truly empty/unstyled cells
                style = _cell_style(c)
                cells.append(Cell(
                    row=c.row, col=c.column, value=c.value,
                    display=format_cell_value(c.value, c.number_format),
                    style=style,
                    hyperlink=c.hyperlink.target if c.hyperlink else None,
                    note=c.comment.text if c.comment else None,
                ))
        merged = [MergedRange(r.min_row, r.min_col, r.max_row, r.max_col)
                  for r in ws.merged_cells.ranges]
        col_widths = {}
        for letter, dim in ws.column_dimensions.items():
            if dim.width:
                try:
                    col_widths[openpyxl.utils.column_index_from_string(letter)] = dim.width
                except ValueError:
                    pass
        sheets.append(Sheet(
            name=ws.title, index=idx,
            max_row=ws.max_row, max_col=ws.max_column,
            cells=cells, merged=merged, col_widths=col_widths,
        ))
    return Workbook(source=xlsx_path, sheets=sheets, media=[])


def extract_media(xlsx_path: str, images_dir: str) -> list[ImageAsset]:
    os.makedirs(images_dir, exist_ok=True)
    assets: list[ImageAsset] = []
    with zipfile.ZipFile(xlsx_path) as z:
        for name in z.namelist():
            if name.startswith("xl/media/"):
                base = os.path.basename(name)
                dest = os.path.join(images_dir, base)
                with z.open(name) as src, open(dest, "wb") as out:
                    shutil.copyfileobj(src, out)
                w = h = None
                try:
                    from PIL import Image
                    with Image.open(dest) as im:
                        w, h = im.size
                except Exception:
                    pass
                assets.append(ImageAsset(
                    id=base, path=os.path.join("images", base),
                    width_px=w, height_px=h, source="media",
                ))
    return assets


def _has_zip_dir(xlsx_path: str, prefix: str) -> bool:
    with zipfile.ZipFile(xlsx_path) as z:
        return any(n.startswith(prefix) for n in z.namelist())


def extract_workbook(xlsx_path: str, out_dir: str, render: bool = True) -> Workbook:
    os.makedirs(out_dir, exist_ok=True)
    images_dir = os.path.join(out_dir, "images")
    shots_dir = os.path.join(out_dir, "screenshots")
    wb = extract_semantic(xlsx_path)
    wb.media = extract_media(xlsx_path, images_dir)
    wb.has_charts = _has_zip_dir(xlsx_path, "xl/charts/")
    wb.has_embeddings = _has_zip_dir(xlsx_path, "xl/embeddings/")
    if render:
        try:
            from .render import render_xlsx_to_images
            shots = render_xlsx_to_images(xlsx_path, shots_dir)
            # Attach all page shots to sheet 0; multi-sheet refinement is a later concern.
            if wb.sheets and shots:
                wb.sheets[0].screenshots = [os.path.join("screenshots", os.path.basename(s))
                                            for s in shots]
        except Exception:
            pass  # rendering is best-effort; absence of soffice must not break extraction
    with open(os.path.join(out_dir, "workbook.json"), "w") as f:
        f.write(wb.to_json())
    return wb
